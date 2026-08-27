"""
Leitura do relatorio de transacoes exportado do Backoffice.

O layout tem um bloco de filtros no topo (quantidade variavel de linhas),
depois a linha "Criado por usuario ...", a linha de cabecalho comecando com
"ID Transacao" e por fim as transacoes.

Nada aqui e fixado por numero de linha: o cabecalho e localizado pelo texto
(RF-05), entao relatorios com mais ou menos filtros funcionam igual — a
amostra de 25/08 tem 30 linhas antes da tabela.

Nem todo relatorio traz o mesmo conjunto de colunas. Coluna ausente nao
quebra a carga: o campo sai vazio e o laudo imprime "-" (RF-09).

DESEMPENHO — a planilha e lida em UMA passada com iter_rows. Em modo
read_only o openpyxl re-varre a planilha a cada `ws.cell(r, c)`, o que torna
o acesso celula a celula quadratico: a leitura das 3.419 linhas da amostra
levava 80 minutos assim, contra menos de um segundo por streaming.
"""
from __future__ import annotations

import re
from datetime import date, datetime, time

from openpyxl import load_workbook

from . import parametros as P
from .dominio import Relatorio, Transito, normalizar

#: campos sem os quais nao da para trabalhar
ESSENCIAIS = ["id"]
#: campos que o laudo imprime — a falta deles e avisada, mas nao impede
DO_LAUDO = ["placa", "categoria", "velocidade"]

ROTULO_AMIGAVEL = {"placa": "Placa", "categoria": "Categoria",
                   "velocidade": "Velocidade"}


class RelatorioInvalido(Exception):
    pass


def _celula(linha: tuple, i: int):
    """Valor da coluna i (1-based), tolerando linhas mais curtas."""
    return linha[i - 1] if 0 < i <= len(linha) else None


# ───────────────────────────────────────────────────────────── cabecalho
def _achar_cabecalho(linhas: list[tuple]) -> int:
    """Indice (0-based) da linha de cabecalho."""
    limite = min(P.LIMITE_BUSCA_CABECALHO, len(linhas))
    for i in range(limite):
        if normalizar(_celula(linhas[i], 1)).startswith(P.MARCA_CABECALHO):
            return i
    raise RelatorioInvalido(
        'Não encontrei a coluna "ID Transação" neste arquivo. '
        "Confirme que é o relatório de transações exportado do Backoffice, "
        "salvo em .xlsx."
    )


def _nomes_das_colunas(linhas: list[tuple], i_cab: int) -> list[tuple[str, str]]:
    """
    Devolve, por coluna, (nome_composto, nome_puro).

    A linha logo acima do cabecalho pode trazer grupos que qualificam colunas
    repetidas — "Categoria" + "Arr." vira "Categoria - Arr." (RF-06). Quando
    essa linha esta vazia, como no relatorio de 25/08, os dois nomes coincidem.
    """
    cab = linhas[i_cab]
    largura = len(cab)
    grupos: dict[int, str] = {}
    if i_cab >= 1:
        atual = ""
        acima = linhas[i_cab - 1]
        for c in range(1, largura + 1):
            bruto = _celula(acima, c)
            if bruto is not None:
                atual = str(bruto).strip()
            grupos[c] = atual

    nomes: list[tuple[str, str]] = []
    for c in range(1, largura + 1):
        bruto = _celula(cab, c)
        puro = str(bruto).strip() if bruto is not None else ""
        grupo = grupos.get(c, "")
        composto = (f"{grupo} - {puro}"
                    if (grupo and grupo != puro and c > 1 and puro) else puro)
        nomes.append((composto, puro))
    return nomes


def _mapear_colunas(nomes: list[tuple[str, str]]) -> dict[str, int]:
    """chave interna -> indice da coluna (1-based). O primeiro que casar vence."""
    mapa: dict[str, int] = {}
    for chave, aliases in P.COLUNAS.items():
        for i, (composto, puro) in enumerate(nomes, start=1):
            if chave in mapa:
                break
            for candidato in (composto, puro):
                if normalizar(candidato) in aliases:
                    mapa[chave] = i
                    break
    return mapa


# ───────────────────────────────────────────────────────────── metadados
def _ler_metadados(linhas: list[tuple], i_cab: int) -> dict[str, str]:
    meta: dict[str, str] = {}
    for i in range(i_cab):
        rotulo = _celula(linhas[i], 1)
        if rotulo is None:
            continue
        rotulo = str(rotulo).strip()

        if rotulo.endswith(":"):
            valor = ""
            for c in range(2, 8):
                bruto = _celula(linhas[i], c)
                if bruto not in (None, ""):
                    valor = str(bruto).strip()
                    break
            meta[rotulo.rstrip(":").strip()] = valor
        elif normalizar(rotulo).startswith("criado por"):
            meta["_origem"] = rotulo

    return meta


def linha_analise(metadados: dict[str, str]) -> str:
    """
    RF-40 — a linha de origem do laudo.

    "Criado por usuário 00598 em 25/08/2026 12:00:02 no nó XXX000 HOST"
                     vira
    "Analisado por usuário 00598 em 25/08/2026 12:00:02"
    """
    bruto = metadados.get("_origem", "").strip()
    if not bruto:
        return f"Analisado em {datetime.now():%d/%m/%Y %H:%M:%S}"
    sem_no = re.sub(r"\s+no\s+.*HOST\s*$", "", bruto, flags=re.IGNORECASE)
    return re.sub(r"^criado\s+por", "Analisado por", sem_no, flags=re.IGNORECASE)


# ──────────────────────────────────────────────────────────────── datas
def _para_datahora(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime.combine(v, time.min)
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt)
            except ValueError:
                continue
    return None


# ───────────────────────────────────────────────────────────────── ler
def ler(caminho: str) -> Relatorio:
    """Le o .xlsx e devolve o Relatorio com todos os transitos."""
    try:
        wb = load_workbook(caminho, data_only=True, read_only=True)
    except Exception as e:                                    # noqa: BLE001
        raise RelatorioInvalido(
            f"Não consegui abrir a planilha: {e}. "
            "Confirme que o arquivo é um .xlsx e não está aberto no Excel."
        ) from e

    try:
        ws = wb[wb.sheetnames[0]]
        linhas = [tuple(l) for l in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if not linhas:
        raise RelatorioInvalido("A planilha está vazia.")

    i_cab = _achar_cabecalho(linhas)
    nomes = _nomes_das_colunas(linhas, i_cab)
    mapa = _mapear_colunas(nomes)

    if any(c not in mapa for c in ESSENCIAIS):
        raise RelatorioInvalido(
            "O arquivo tem o cabeçalho no lugar certo, mas falta a coluna "
            "ID Transação."
        )

    rel = Relatorio(
        arquivo=caminho,
        metadados=_ler_metadados(linhas, i_cab),
        colunas_achadas={k: nomes[i - 1][0] for k, i in mapa.items()},
        colunas_faltando=[c for c in DO_LAUDO if c not in mapa],
    )

    col_id = mapa["id"]
    for n, linha in enumerate(linhas[i_cab + 1:], start=i_cab + 2):
        bruto = _celula(linha, col_id)
        texto_id = "" if bruto is None else str(bruto).strip()
        if not texto_id:
            break
        if normalizar(texto_id).startswith(P.MARCA_FIM):
            break

        campos = {}
        for chave, indice in mapa.items():
            valor = _celula(linha, indice)
            if chave == "data_hora":
                valor = _para_datahora(valor)
            campos[chave] = valor

        rel.transitos.append(Transito(id=texto_id, linha=n, campos=campos))

    if not rel.transitos:
        raise RelatorioInvalido(
            "O arquivo tem o cabeçalho certo, mas nenhuma transação abaixo dele."
        )

    return rel


def resumo_da_carga(rel: Relatorio) -> dict:
    """O que a tela de carga mostra antes de a conferência começar (RF-09)."""
    return {
        "arquivo": rel.arquivo,
        "total": len(rel.transitos),
        "concessionaria": rel.concessionaria,
        "periodo": rel.periodo,
        "pracas": rel.pracas,
        "pistas": rel.pistas,
        "colunas_faltando": [ROTULO_AMIGAVEL.get(c, c)
                             for c in rel.colunas_faltando],
        "contramao_sinalizados": sum(1 for t in rel.transitos if t.eh_contramao),
    }
